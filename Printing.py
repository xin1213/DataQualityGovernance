import re
import os
import openpyxl


def create_directory(path, project_name="DataQualityGovernance"):
    absolute_path = os.path.abspath(path)

    reg_ex_1 = re.compile('\\\\' + project_name + '\\\\')
    root_directory = reg_ex_1.split(absolute_path)[0]
    follow_directory = reg_ex_1.split(absolute_path)[1]

    reg_ex_2 = re.compile('\\\\')
    directories_to_check = reg_ex_2.split(follow_directory)[:-1]

    directories_path = root_directory + '\\\\' + project_name
    for x in directories_to_check:
        directories_path = directories_path + "\\" + x
        if not os.path.exists(directories_path):
            os.makedirs(directories_path)
    print("Create directory successfully!")


def create_workbook(path):
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        wb.save(filename=path)
        wb.close()
        print("Create workbook successfully!")


def save_to_excel_1d(data, column, wb_name, sheet_name, start_column, start_row):
    wb = openpyxl.load_workbook(filename=wb_name)
    try:
        sheet = wb[sheet_name]
    except:
        wb.create_sheet(title=sheet_name)
    finally:
        sheet = wb[sheet_name]
    _ = sheet.cell(row=1, column=start_column, value=str(column))
    for r in range(start_row, len(data) + start_row):
        _ = sheet.cell(row=r, column=start_column, value=str(data[r - start_row]))
    wb.save(filename=wb_name)
    wb.close()
    print("Save the file successfully!")


def save_to_excel_2d(data, columns, wb_name, sheet_name, start_column, start_row):
    wb = openpyxl.load_workbook(filename=wb_name)
    try:
        sheet = wb[sheet_name]
    except:
        wb.create_sheet(title=sheet_name)
    finally:
        sheet = wb[sheet_name]
    for field in range(start_column, len(columns) + start_column):
        _ = sheet.cell(row=1, column=field, value=str(columns[field - start_column]))
        for r in range(start_row, len(data) + start_row):
            _ = sheet.cell(row=r, column=field, value=str(data[r - start_row][field - start_column]))
    wb.save(filename=wb_name)
    wb.close()
    print("Save the file successfully!")